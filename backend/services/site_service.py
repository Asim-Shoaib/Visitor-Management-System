from typing import Optional, Dict, List
from datetime import datetime, timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from backend.database.connection import Database

db = Database()


def get_all_sites() -> List[Dict]:
    rows = db.fetchall("SELECT site_id, site_name, address FROM sites ORDER BY site_name")
    return rows if rows is not None else []


def create_site(site_name: str, address: Optional[str] = None, created_by_user_id: int = None) -> Optional[int]:
    """
    Create a new site record.
    Returns site_id on success, None on failure.
    """
    from backend.utils.validator import validate_name
    from backend.utils.db_logger import log_action
    
    # Validate inputs
    if not validate_name(site_name):
        return None
    
    # Check if site already exists
    existing = db.fetchone("SELECT site_id FROM sites WHERE site_name = %s", (site_name.strip(),))
    if existing:
        return None
    
    # Insert site
    insert_sql = """
        INSERT INTO sites (site_name, address)
        VALUES (%s, %s)
    """
    success = db.execute(insert_sql, (site_name.strip(), address.strip() if address else None))
    
    if success:
        # Get the inserted site_id
        site = db.fetchone("SELECT site_id FROM sites WHERE site_name = %s ORDER BY site_id DESC LIMIT 1", (site_name.strip(),))
        if site:
            site_id = site['site_id']
            if created_by_user_id:
                log_action(created_by_user_id, "create_site", f"Created site {site_name} (ID: {site_id}, Address: {address})")
            return site_id
    
    return None


def get_all_departments() -> List[Dict]:
    rows = db.fetchall("SELECT department_id, name FROM departments ORDER BY name")
    return rows if rows is not None else []


def get_all_employees() -> List[Dict]:
    rows = db.fetchall("""
        SELECT 
            e.employee_id, 
            e.name, 
            e.hourly_rate,
            d.name as department_name
        FROM employees e
        LEFT JOIN departments d ON e.department_id = d.department_id
        ORDER BY e.name
    """)
    return rows if rows is not None else []


def get_active_employees_count() -> int:
    """Get count of employees who are currently signed in"""
    # An employee is signed in if their last scan was signin
    result = db.fetchone("""
        SELECT COUNT(DISTINCT esl.emp_qr_id) as active_count
        FROM employeescanlogs esl
        INNER JOIN (
            SELECT emp_qr_id, MAX(timestamp) as max_timestamp
            FROM employeescanlogs
            GROUP BY emp_qr_id
        ) latest ON esl.emp_qr_id = latest.emp_qr_id 
            AND esl.timestamp = latest.max_timestamp
        WHERE esl.scan_status = 'signin'
    """)
    return result['active_count'] if result else 0


def get_signed_in_employees() -> List[Dict]:
    """Get list of currently signed-in employees"""
    rows = db.fetchall("""
        SELECT DISTINCT
            e.employee_id,
            e.name,
            e.hourly_rate,
            d.name as department_name,
            esl.timestamp as last_scan_time
        FROM employees e
        INNER JOIN employeeqrcodes eqr ON e.employee_id = eqr.employee_id
        INNER JOIN employeescanlogs esl ON eqr.emp_qr_id = esl.emp_qr_id
        LEFT JOIN departments d ON e.department_id = d.department_id
        INNER JOIN (
            SELECT emp_qr_id, MAX(timestamp) as max_timestamp
            FROM employeescanlogs
            GROUP BY emp_qr_id
        ) latest ON esl.emp_qr_id = latest.emp_qr_id 
            AND esl.timestamp = latest.max_timestamp
        WHERE esl.scan_status = 'signin'
        ORDER BY esl.timestamp DESC
    """)
    return rows if rows is not None else []


def get_employee_status(employee_id: int) -> Optional[Dict]:
    """Get current status of an employee (signed in/out)"""
    employee = db.fetchone("""
        SELECT e.employee_id, e.name, e.hourly_rate, d.name as department_name
        FROM employees e
        LEFT JOIN departments d ON e.department_id = d.department_id
        WHERE e.employee_id = %s
    """, (employee_id,))
    
    if not employee:
        return None
    
    # Get last scan
    last_scan = db.fetchone("""
        SELECT esl.scan_status, esl.timestamp
        FROM employeeqrcodes eqr
        INNER JOIN employeescanlogs esl ON eqr.emp_qr_id = esl.emp_qr_id
        WHERE eqr.employee_id = %s
        ORDER BY esl.timestamp DESC
        LIMIT 1
    """, (employee_id,))
    
    is_signed_in = last_scan and last_scan['scan_status'] == 'signin' if last_scan else False
    
    return {
        'employee_id': employee['employee_id'],
        'name': employee['name'],
        'department': employee['department_name'],
        'hourly_rate': float(employee['hourly_rate']) if employee['hourly_rate'] else 0.0,
        'is_signed_in': is_signed_in,
        'last_scan_time': last_scan['timestamp'].isoformat() if last_scan else None,
        'last_scan_status': last_scan['scan_status'] if last_scan else None
    }


def get_employee_logs(employee_id: int, days: int = 30) -> List[Dict]:
    """Get employee logs for last N days"""
    rows = db.fetchall("""
        SELECT 
            esl.scan_id,
            esl.scan_status,
            esl.timestamp,
            eqr.code_value as qr_code_value
        FROM employeescanlogs esl
        INNER JOIN employeeqrcodes eqr ON esl.emp_qr_id = eqr.emp_qr_id
        WHERE eqr.employee_id = %s
          AND DATE(esl.timestamp) >= DATE_SUB(CURDATE(), INTERVAL %s DAY)
        ORDER BY esl.timestamp DESC
    """, (employee_id, days))
    return rows if rows is not None else []


def create_employee(name: str, hourly_rate: float, department_id: int, created_by_user_id: int) -> Optional[int]:
    """
    Create a new employee record.
    Returns employee_id on success, None on failure.
    """
    from backend.utils.validator import validate_name
    from backend.utils.db_logger import log_action
    
    # Validate inputs
    if not validate_name(name):
        return None
    
    if not isinstance(hourly_rate, (int, float)) or hourly_rate < 0:
        return None
    
    # Validate department exists
    department = db.fetchone("SELECT department_id FROM departments WHERE department_id = %s", (department_id,))
    if not department:
        return None
    
    # Insert employee
    insert_sql = """
        INSERT INTO employees (name, hourly_rate, department_id)
        VALUES (%s, %s, %s)
    """
    success = db.execute(insert_sql, (name.strip(), float(hourly_rate), department_id))
    
    if success:
        # Get the inserted employee_id
        employee = db.fetchone("SELECT employee_id FROM employees WHERE name = %s AND department_id = %s ORDER BY employee_id DESC LIMIT 1", (name.strip(), department_id))
        if employee:
            employee_id = employee['employee_id']
            log_action(created_by_user_id, "create_employee", f"Created employee {name} (ID: {employee_id}, Rate: {hourly_rate}, Dept: {department_id})")
            return employee_id
    
    return None


def calculate_employee_salary(employee_id: int, start_date: Optional[str] = None, end_date: Optional[str] = None) -> Optional[Dict]:
    """
    Calculate employee salary based on sign-in/sign-out logs for a date range.
    
    Args:
        employee_id: Employee ID
        start_date: Start date (YYYY-MM-DD format), defaults to 30 days ago
        end_date: End date (YYYY-MM-DD format), defaults to today
    
    Returns:
        Dict with total_hours, total_days, salary, and detailed breakdown, or None if employee not found
    """
    # Get employee info
    employee = db.fetchone("""
        SELECT e.employee_id, e.name, e.hourly_rate, d.name as department_name
        FROM employees e
        LEFT JOIN departments d ON e.department_id = d.department_id
        WHERE e.employee_id = %s
    """, (employee_id,))
    
    if not employee:
        return None
    
    hourly_rate = float(employee["hourly_rate"]) if employee["hourly_rate"] else 0.0
    
    # Set default date range if not provided
    if not end_date:
        end_date = datetime.now().strftime('%Y-%m-%d')
    if not start_date:
        start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    
    # Parse dates
    try:
        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
        end_dt = datetime.strptime(end_date, '%Y-%m-%d')
        # Include full end date (end of day)
        end_dt = end_dt.replace(hour=23, minute=59, second=59)
    except ValueError:
        return None
    
    # Get all scans in the date range
    scans = db.fetchall("""
        SELECT esl.scan_status, esl.timestamp
        FROM employeescanlogs esl
        JOIN employeeqrcodes eqr ON esl.emp_qr_id = eqr.emp_qr_id
        WHERE eqr.employee_id = %s
          AND esl.timestamp >= %s
          AND esl.timestamp <= %s
        ORDER BY esl.timestamp ASC
    """, (employee_id, start_dt, end_dt))
    
    total_hours = 0.0
    signin_time = None
    days_worked = set()
    daily_hours = []  # List of (date, hours) tuples
    
    for scan in scans:
        scan_time = scan["timestamp"]
        scan_date = scan_time.date()
        
        if scan["scan_status"] == "signin":
            signin_time = scan_time
            days_worked.add(scan_date)
        elif scan["scan_status"] == "signout" and signin_time:
            checkout_time = scan_time
            hours = (checkout_time - signin_time).total_seconds() / 3600.0
            hours = max(0, hours)  # Ensure non-negative
            total_hours += hours
            daily_hours.append({
                "date": scan_date.strftime('%Y-%m-%d'),
                "signin": signin_time.strftime('%Y-%m-%d %H:%M:%S'),
                "signout": checkout_time.strftime('%Y-%m-%d %H:%M:%S'),
                "hours": round(hours, 2)
            })
            signin_time = None
    
    # Handle case where employee signed in but didn't sign out (count until end of day)
    if signin_time:
        end_of_day = signin_time.replace(hour=23, minute=59, second=59)
        if end_of_day > signin_time:
            hours = (end_of_day - signin_time).total_seconds() / 3600.0
            hours = max(0, hours)
            total_hours += hours
            daily_hours.append({
                "date": signin_time.date().strftime('%Y-%m-%d'),
                "signin": signin_time.strftime('%Y-%m-%d %H:%M:%S'),
                "signout": end_of_day.strftime('%Y-%m-%d %H:%M:%S'),
                "hours": round(hours, 2),
                "incomplete": True
            })
    
    salary = total_hours * hourly_rate
    
    return {
        "employee_id": employee_id,
        "employee_name": employee["name"],
        "department": employee["department_name"],
        "hourly_rate": hourly_rate,
        "start_date": start_date,
        "end_date": end_date,
        "total_days_worked": len(days_worked),
        "total_hours": round(total_hours, 2),
        "salary": round(salary, 2),
        "daily_breakdown": daily_hours
    }


def export_salary_report_to_excel(employee_id: int, start_date: Optional[str] = None, end_date: Optional[str] = None) -> Optional[BytesIO]:
    """
    Export employee salary report to Excel format.
    
    Args:
        employee_id: Employee ID
        start_date: Start date (YYYY-MM-DD format)
        end_date: End date (YYYY-MM-DD format)
    
    Returns:
        BytesIO object containing the Excel file, or None if employee not found
    """
    salary_data = calculate_employee_salary(employee_id, start_date, end_date)
    
    if not salary_data:
        return None
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Salary Report"
    
    # Define header styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Write summary section
    row = 1
    ws.cell(row=row, column=1, value="Employee Salary Report").font = Font(bold=True, size=14)
    row += 2
    
    summary_data = [
        ["Employee Name:", salary_data["employee_name"]],
        ["Department:", salary_data["department"] or "N/A"],
        ["Employee ID:", salary_data["employee_id"]],
        ["Hourly Rate:", f"${salary_data['hourly_rate']:.2f}"],
        ["Period:", f"{salary_data['start_date']} to {salary_data['end_date']}"],
        ["Total Days Worked:", salary_data["total_days_worked"]],
        ["Total Hours:", f"{salary_data['total_hours']:.2f}"],
        ["Total Salary:", f"${salary_data['salary']:.2f}"]
    ]
    
    for label, value in summary_data:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1
    
    row += 2
    
    # Write daily breakdown header
    headers = ["Date", "Sign-In Time", "Sign-Out Time", "Hours Worked", "Notes"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1
    
    # Write daily breakdown data
    for day in salary_data["daily_breakdown"]:
        ws.cell(row=row, column=1, value=day["date"])
        ws.cell(row=row, column=2, value=day["signin"])
        ws.cell(row=row, column=3, value=day["signout"])
        ws.cell(row=row, column=4, value=day["hours"])
        if day.get("incomplete"):
            ws.cell(row=row, column=5, value="Incomplete (no sign-out recorded)")
        row += 1
    
    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = len(headers[col_num - 1])
        
        # Find max length in column
        for row_num in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row_num, column=col_num).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        
        # Set column width (with some padding)
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file

