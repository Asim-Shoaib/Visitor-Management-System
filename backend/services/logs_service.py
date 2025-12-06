from datetime import datetime
from typing import List, Dict, Optional
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from backend.database.connection import Database

db = Database()


def get_access_logs(start_date: Optional[str] = None, end_date: Optional[str] = None, action: Optional[str] = None) -> List[Dict]:
    """
    Get access logs from AccessLogs table with optional filters.
    
    Args:
        start_date: Start date filter (YYYY-MM-DD format)
        end_date: End date filter (YYYY-MM-DD format)
        action: Action type filter (e.g., 'login', 'visitor_checkin')
    
    Returns:
        List of log records with user and action information
    """
    base_sql = """
        SELECT 
            al.log_id,
            al.user_id,
            u.username,
            al.action,
            al.details,
            al.timestamp
        FROM AccessLogs al
        JOIN Users u ON al.user_id = u.user_id
        WHERE 1=1
    """
    
    params = []
    
    if start_date:
        base_sql += " AND DATE(al.timestamp) >= %s"
        params.append(start_date)
    
    if end_date:
        base_sql += " AND DATE(al.timestamp) <= %s"
        params.append(end_date)
    
    if action:
        base_sql += " AND al.action = %s"
        params.append(action)
    
    base_sql += " ORDER BY al.timestamp DESC LIMIT 1000"
    
    logs = db.fetchall(base_sql, tuple(params))
    return logs


def export_access_logs_to_excel(start_date: Optional[str] = None, end_date: Optional[str] = None, action: Optional[str] = None) -> BytesIO:
    """
    Export access logs to Excel format.
    
    Args:
        start_date: Start date filter (YYYY-MM-DD format)
        end_date: End date filter (YYYY-MM-DD format)
        action: Action type filter (e.g., 'login', 'visitor_checkin')
    
    Returns:
        BytesIO object containing the Excel file
    """
    # Get logs using the same query logic
    logs = get_access_logs(start_date, end_date, action)
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Access Logs"
    
    # Define header row
    headers = ["Log ID", "User ID", "Username", "Action", "Details", "Timestamp"]
    
    # Style for header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data rows
    for row_num, log in enumerate(logs, 2):
        ws.cell(row=row_num, column=1, value=log.get('log_id', ''))
        ws.cell(row=row_num, column=2, value=log.get('user_id', ''))
        ws.cell(row=row_num, column=3, value=log.get('username', ''))
        ws.cell(row=row_num, column=4, value=log.get('action', ''))
        ws.cell(row=row_num, column=5, value=log.get('details', ''))
        
        # Format timestamp
        timestamp = log.get('timestamp')
        if timestamp:
            if isinstance(timestamp, str):
                timestamp = datetime.fromisoformat(timestamp.replace('Z', '+00:00'))
            elif isinstance(timestamp, datetime):
                pass
            else:
                timestamp = str(timestamp)
            ws.cell(row=row_num, column=6, value=timestamp)
        else:
            ws.cell(row=row_num, column=6, value='')
    
    # Auto-adjust column widths
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        max_length = len(header)
        
        # Find max length in column
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
            if row[0].value:
                max_length = max(max_length, len(str(row[0].value)))
        
        # Set column width (with some padding)
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file

